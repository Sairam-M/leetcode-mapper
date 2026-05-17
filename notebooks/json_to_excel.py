import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

COLUMNS = [
    "Pattern Title",
    "Segment",
    "Link Type",
    "Source URL",
    "Mapping Success",
    "LLM Suggested Title",
    "LC Title",
    "LC Link",
    "Difficulty",
    "Acceptance Rate",
    "Topics",
    "Pattern",
    "Core Constraint",
    "Why it Matches",
    "Key Difference",
    "Confidence",
]

COLUMN_WIDTHS = {
    "Pattern Title": 30,
    "Segment": 10,
    "Link Type": 15,
    "Source URL": 40,
    "Mapping Success": 15,
    "LLM Suggested Title": 30,
    "LC Title": 30,
    "LC Link": 40,
    "Difficulty": 12,
    "Acceptance Rate": 15,
    "Topics": 25,
    "Pattern": 20,
    "Core Constraint": 25,
    "Why it Matches": 50,
    "Key Difference": 40,
    "Confidence": 15,
}

HEADER_FILL = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)

CONFIDENCE_FILLS = {
    "Strong match": PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE"),
    "Moderate match": PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C"),
    "Weak match": PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE"),
}

TYPE_FILLS = {
    "gfg_problem": PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF"),
    "gfg_topical": PatternFill("solid", start_color="EDEDED", end_color="EDEDED"),
    "leetcode_problem": PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB"),
    "leetcode_list": PatternFill("solid", start_color="EBF3FB", end_color="EBF3FB"),
    "notion": PatternFill("solid", start_color="F5F0FF", end_color="F5F0FF"),
    "other": PatternFill("solid", start_color="FFF8ED", end_color="FFF8ED"),
}

thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def build_rows(entry):
    title = entry.get("title", "")
    segment = entry.get("segment", "")
    classified = {c["url"]: c["type"] for c in entry.get("classified_links", [])}

    mapped_by_url = {}
    for link in entry.get("mapped_links", []):
        mapped_by_url[link["url"]] = link

    rows = []
    for url, link_type in classified.items():
        mapped = mapped_by_url.get(url, {})
        result = mapped.get("mapped_result") or {}
        success = mapped.get("mapping_success")

        if success is True:
            success_str = "Yes"
        elif success is False:
            success_str = "No"
        else:
            success_str = "N/A"

        row = {
            "Pattern Title": title,
            "Segment": segment,
            "Link Type": link_type,
            "Source URL": url,
            "Mapping Success": success_str,
            "LLM Suggested Title": result.get("title", ""),
            "LC Title": result.get("leetcode_title", ""),
            "LC Link": result.get("leetcode_link", ""),
            "Difficulty": result.get("difficulty", ""),
            "Acceptance Rate": result.get("acceptance_rate", ""),
            "Topics": result.get("topics", ""),
            "Pattern": result.get("pattern", ""),
            "Core Constraint": result.get("core_constraint", ""),
            "Why it Matches": result.get("why_it_matches", ""),
            "Key Difference": result.get("key_difference", ""),
            "Confidence": result.get("confidence", ""),
        }
        rows.append((link_type, row))

    return rows


def write_sheet(wb, sheet_name, entries):
    ws = wb.create_sheet(title=sheet_name)

    # Header row
    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_name]

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Sort entries by segment
    sorted_entries = sorted(entries, key=lambda e: (
        int(e.get("segment") or 0),
        e.get("title", "")
    ))

    current_row = 2
    for entry in sorted_entries:
        rows = build_rows(entry)
        for link_type, row_data in rows:
            for col_idx, col_name in enumerate(COLUMNS, 1):
                cell = ws.cell(row=current_row, column=col_idx, value=row_data.get(col_name, ""))
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = thin_border

                # Row fill based on confidence or link type
                confidence = row_data.get("Confidence", "")
                if confidence in CONFIDENCE_FILLS:
                    cell.fill = CONFIDENCE_FILLS[confidence]
                else:
                    cell.fill = TYPE_FILLS.get(link_type, PatternFill())

                # Hyperlinks for URLs
                if col_name in ("Source URL", "LC Link") and row_data.get(col_name, "").startswith("http"):
                    cell.hyperlink = row_data[col_name]
                    cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")

            ws.row_dimensions[current_row].height = 40
            current_row += 1

    return ws


def json_to_excel(input_path, output_path):
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    # Group entries by module
    modules = defaultdict(list)
    for entry in data:
        module = entry.get("module") or "Unknown"
        modules[module].append(entry)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for module_num in sorted(modules.keys(), key=lambda x: int(x) if str(x).isdigit() else 999):
        sheet_name = f"Module {module_num}"
        write_sheet(wb, sheet_name, modules[module_num])

    wb.save(output_path)
    print(f"Saved: {output_path}")
    print(f"Modules: {sorted(modules.keys(), key=lambda x: int(x) if str(x).isdigit() else 999)}")
    print(f"Total entries: {len(data)}")


if __name__ == "__main__":
    json_to_excel("./data-local/mapped_links.json", "./data-local/gfg_lc_guide.xlsx")