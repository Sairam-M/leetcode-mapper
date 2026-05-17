import json
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

MANUAL_EXCEL = "./data-local/Leetcode Equivalent.xlsx"
GUIDE_EXCEL = "./data-local/gfg_lc_guide.xlsx"
OUTPUT_EXCEL = "./data-local/gfg_lc_guide_verified.xlsx"


# ── 1. Load manual Excel ground truth ────────────────────────────────────────

def strip_number_prefix(title):
    if not isinstance(title, str):
        return ''
    return re.sub(r'^\d+\.\s*', '', title.strip()).strip()

def extract_lc_titles(raw):
    if not isinstance(raw, str) or not raw.strip():
        return []
    skip_prefixes = ['no direct', 'edu -', 'sorting-related', 'educational']
    if any(raw.lower().startswith(p) for p in skip_prefixes):
        return []
    parts = re.split(r'\s+(?=\d+\.)', raw)
    titles = [strip_number_prefix(p) for p in parts if p.strip()]
    return [t for t in titles if t]

def load_manual_excel(path):
    xl = pd.read_excel(path, sheet_name=None, usecols="A:E")
    ground_truth = {}
    for sheet_name, df in xl.items():
        for _, row in df.iterrows():
            url = str(row.get('Original Links', '') or '').strip()
            if not url.startswith('https://www.geeksforgeeks.org'):
                continue
            lc_titles = extract_lc_titles(row.get('LeetCode Equivalent(s)', ''))
            verified = str(row.get('Verified', '') or '').strip().lower() == 'verified'
            if url not in ground_truth:
                ground_truth[url] = {'lc_titles': set(), 'verified': False}
            ground_truth[url]['lc_titles'].update(lc_titles)
            if verified:
                ground_truth[url]['verified'] = True
    return ground_truth


# ── 2. Determine verification status ─────────────────────────────────────────

def get_verification(url, mapped_lc_title, link_type, ground_truth):
    if link_type != 'gfg_problem':
        return 'Not Applicable'
    if url not in ground_truth:
        return 'Unverified'
    gt = ground_truth[url]
    if not gt['verified']:
        return 'Unverified'
    if not mapped_lc_title:
        return 'Unverified'
    match = any(
        mapped_lc_title.strip().lower() == t.lower()
        for t in gt['lc_titles']
    )
    return 'Verified - Correct' if match else 'Verified - Incorrect'


# ── 3. Add Verified column to guide Excel ─────────────────────────────────────

VERIFICATION_FILLS = {
    'Verified - Correct':   PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE"),
    'Verified - Incorrect': PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE"),
    'Unverified':           PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C"),
    'Not Applicable':       PatternFill("solid", start_color="EDEDED", end_color="EDEDED"),
}

def add_verification_to_guide(guide_path, output_path, ground_truth):
    wb = openpyxl.load_workbook(guide_path)
    all_stats = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = {cell.value: cell.column for cell in ws[1] if cell.value}
        source_url_col = headers.get('Source URL')
        link_type_col  = headers.get('Link Type')
        lc_title_col   = headers.get('LC Title')
        confidence_col = headers.get('Confidence')
        difficulty_col = headers.get('Difficulty')
        pattern_col    = headers.get('Pattern')

        if not all([source_url_col, link_type_col, lc_title_col]):
            continue

        verified_col = ws.max_column + 1
        header_cell = ws.cell(row=1, column=verified_col, value='Verified')
        header_cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        header_cell.fill = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(verified_col)].width = 20

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            url        = str(row[source_url_col - 1].value or '').strip()
            link_type  = str(row[link_type_col - 1].value or '').strip()
            lc_title   = str(row[lc_title_col - 1].value or '').strip()
            confidence = str(row[confidence_col - 1].value or '').strip() if confidence_col else ''
            difficulty = str(row[difficulty_col - 1].value or '').strip() if difficulty_col else ''
            pattern    = str(row[pattern_col - 1].value or '').strip() if pattern_col else ''

            verification = get_verification(url, lc_title, link_type, ground_truth)

            cell = ws.cell(row=row[0].row, column=verified_col, value=verification)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = VERIFICATION_FILLS.get(verification, PatternFill())

            all_stats.append({
                'sheet':        sheet_name,
                'link_type':    link_type,
                'lc_title':     lc_title,
                'verification': verification,
                'url':          url,
                'confidence':   confidence,
                'difficulty':   difficulty,
                'pattern':      pattern,
            })

    wb.save(output_path)
    return all_stats


# ── 4. Print stats ────────────────────────────────────────────────────────────

def pct(n, d):
    return f"{n/d*100:.1f}%" if d > 0 else "N/A"

def print_stats(all_stats, ground_truth):
    gfg_rows   = [s for s in all_stats if s['link_type'] == 'gfg_problem']
    mapped_urls = set(s['url'] for s in gfg_rows)
    gt_urls     = set(ground_truth.keys())

    correct    = [s for s in gfg_rows if s['verification'] == 'Verified - Correct']
    incorrect  = [s for s in gfg_rows if s['verification'] == 'Verified - Incorrect']
    unverified = [s for s in gfg_rows if s['verification'] == 'Unverified']
    verified_rows = correct + incorrect

    coverage     = len(mapped_urls & gt_urls)
    total_gt     = len(gt_urls)
    total_mapped = len(gfg_rows)

    print("=" * 65)
    print("📊 VERIFICATION SUMMARY")
    print("=" * 65)
    print(f"GFG links in manual Excel (ground truth): {total_gt}")
    print(f"GFG problem rows in guide Excel:          {total_mapped}")
    print(f"Overlap (in both):                        {coverage} ({pct(coverage, total_gt)} coverage)")
    print(f"In manual Excel but not in guide:         {total_gt - coverage}")
    print()
    print(f"Verified - Correct:    {len(correct):4d}  ({pct(len(correct), total_mapped)} of all GFG rows)")
    print(f"Verified - Incorrect:  {len(incorrect):4d}  ({pct(len(incorrect), total_mapped)} of all GFG rows)")
    print(f"Unverified:            {len(unverified):4d}  ({pct(len(unverified), total_mapped)} of all GFG rows)")
    print()
    print(f"Overall Accuracy (verified rows only):    {pct(len(correct), len(verified_rows))}")
    print()

    # Accuracy by confidence level
    print("Accuracy by Confidence Level:")
    print(f"  {'Confidence':<20} {'Total':>6} {'Correct':>8} {'Incorrect':>10} {'Accuracy':>10}")
    print(f"  {'-'*56}")
    for conf in ['Strong match', 'Moderate match', 'Weak match']:
        conf_rows = [s for s in verified_rows if s['confidence'] == conf]
        c = len([s for s in conf_rows if s['verification'] == 'Verified - Correct'])
        i = len([s for s in conf_rows if s['verification'] == 'Verified - Incorrect'])
        print(f"  {conf:<20} {len(conf_rows):>6} {c:>8} {i:>10} {pct(c, len(conf_rows)):>10}")
    print()

    # Accuracy by difficulty
    print("Accuracy by Difficulty (verified rows):")
    print(f"  {'Difficulty':<12} {'Total':>6} {'Correct':>8} {'Incorrect':>10} {'Accuracy':>10}")
    print(f"  {'-'*48}")
    for diff in ['Easy', 'Medium', 'Hard']:
        diff_rows = [s for s in verified_rows if s['difficulty'] == diff]
        c = len([s for s in diff_rows if s['verification'] == 'Verified - Correct'])
        i = len([s for s in diff_rows if s['verification'] == 'Verified - Incorrect'])
        print(f"  {diff:<12} {len(diff_rows):>6} {c:>8} {i:>10} {pct(c, len(diff_rows)):>10}")
    print()

    # Breakdown by module
    print("Breakdown by Module:")
    print(f"  {'Module':<12} {'GFG Rows':>10} {'Correct':>8} {'Incorrect':>10} {'Unverified':>12} {'Accuracy':>10}")
    print(f"  {'-'*64}")
    for sheet in sorted(set(s['sheet'] for s in gfg_rows)):
        sheet_rows = [s for s in gfg_rows if s['sheet'] == sheet]
        c = len([s for s in sheet_rows if s['verification'] == 'Verified - Correct'])
        i = len([s for s in sheet_rows if s['verification'] == 'Verified - Incorrect'])
        u = len([s for s in sheet_rows if s['verification'] == 'Unverified'])
        print(f"  {sheet:<12} {len(sheet_rows):>10} {c:>8} {i:>10} {u:>12} {pct(c, c+i):>10}")
    print()

    # Strong match that are incorrect — worth investigating
    strong_incorrect = [s for s in incorrect if s['confidence'] == 'Strong match']
    if strong_incorrect:
        print(f"⚠️  Strong matches that are Incorrect ({len(strong_incorrect)}):")
        for s in strong_incorrect[:5]:
            gt = ground_truth.get(s['url'], {})
            gt_titles = ', '.join(gt.get('lc_titles', set()))
            print(f"  URL:     ...{s['url'][-50:]}")
            print(f"  Mapped:  {s['lc_title']}")
            print(f"  Correct: {gt_titles}")
            print()

    # Sample incorrect
    if incorrect:
        print(f"Sample Incorrect Mappings (first 5):")
        for s in incorrect[:5]:
            gt = ground_truth.get(s['url'], {})
            gt_titles = ', '.join(gt.get('lc_titles', set()))
            print(f"  URL:     ...{s['url'][-50:]}")
            print(f"  Mapped:  {s['lc_title']} ({s['confidence']})")
            print(f"  Correct: {gt_titles}")
            print()

    print("=" * 65)


# ── Main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print("Loading manual Excel ground truth...")
    ground_truth = load_manual_excel(MANUAL_EXCEL)
    print(f"✓ {len(ground_truth)} GFG URLs loaded\n")

    print("Adding Verified column to guide Excel...")
    all_stats = add_verification_to_guide(GUIDE_EXCEL, OUTPUT_EXCEL, ground_truth)
    print(f"✓ Saved to {OUTPUT_EXCEL}\n")

    print_stats(all_stats, ground_truth)